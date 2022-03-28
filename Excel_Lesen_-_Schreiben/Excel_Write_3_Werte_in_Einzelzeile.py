from openpyxl import Workbook
from openpyxl.utils import get_column_letter

myList=['Frank','hatte','gestern','Geburtstag']


def writeInOneLine(listName, excelFileName):
    wb = Workbook()
    ws = wb.active
    spalten=[]
    for x in list(range(1,len(myList)+1)):
        spaltenName = get_column_letter(x)
        spalten.append(spaltenName)
    for el in spalten:
        i=0
        zeilenNummer = "1"
        while i < len(spalten):
            ws[spalten[i]+zeilenNummer]=myList[i]
            i+=1
    wb.save(excelFileName)

writeInOneLine(myList,"Bsp_Excel_Write_3.xlsx")


