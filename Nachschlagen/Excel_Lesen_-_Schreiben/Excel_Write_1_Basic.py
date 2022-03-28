from openpyxl import Workbook

    # zuerst wird eine workbook Instanz erstellt
wb=Workbook()
    # anschliessend die worksheet Instanz
ws=wb.active
    # Jetzt wird geschrieben dabei werden die einzelnen Zellen direkt angesprochen
ws['A1']="Hello"
ws['A2']="World"
ws['B1']="Frank"
ws['B2']="Menzel"
    # Das workbook wird gespeichert
wb.save("Bsp_Excel_Write.xlsx")
