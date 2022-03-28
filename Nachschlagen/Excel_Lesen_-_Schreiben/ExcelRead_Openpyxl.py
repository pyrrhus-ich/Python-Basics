from openpyxl import load_workbook

    #Basis für Nutzung von openpyxl 
    # 1. Workbook erstellen => hier "wb"
wb=load_workbook(filename="Test.xlsx")
    # 2. Worksheet erstellen => hier ws
ws=wb["Tabelle1"]
    # Iterator erstellen
def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]
print(list(iter_rows(ws)))
    # Mit Hilfe des Iterators alle Spalten des Excel Files in einer
    # Liste von Listen speichern
val = []
for el in iter_rows(ws):
    val.append(el)
print(val)
    # Mit Hilfe des Iterators die Werte einer Spalte in einer 
    # Liste speichern
val0=[]
for el in iter_rows(ws):
    val0.append(el[0])
print(val0)
print(val0[0])




