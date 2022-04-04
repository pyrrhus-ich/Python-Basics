import os
from pathlib import Path
from openpyxl import Workbook, load_workbook

    # wird für srcFile verwendet 
def by_mtime(file):
    return file.stat().st_mtime

workDir = os.getcwd()           
srcFld = workDir + "\\srcZir\\"  
srcFile = max((p for p in Path(srcFld).iterdir() if p.is_file()), key=by_mtime) #ermittelt immer die neueste Datei im Ordner
srcFileName=os.path.split(srcFile)[1] # Nur damit das dstFile diesen Namen verwenden kann
dstFile = workDir +"\\dstZir\\changed-"+ srcFileName
# Wichtig für das korrekte Einlesen des SourceFiles
wb=load_workbook(filename=srcFile)
ws=wb.worksheets[0]

"""
Wichtig die folgenden Variablen bezeichnet den Index der Spalte in der das "Date of creation" und die Zir Reparatur
auftragsnummer ("M199-12345-R") steht
Index fängt bei 0 an zu zählen. Also ist 2 die 3. Spalte
"""
indRepairNumber= 1
indDateOfCreation = 3

# Variable die alls Werte aus dem sourceFile aufnimmt
val = []

# Wichtig für das erstellen des neuen Files
wbDst = Workbook()       # erzeugt Workbook Objekt mit einem Sheet
wsDst = wbDst.active     # aktiviert das erste sheet
wsDst.title="ZIR_ALL"
wsDst.sheet_properties.tabColor = "FF8800"

