from ctypes import alignment
from tkinter import CENTER, HORIZONTAL
from openpyxl.styles import PatternFill, Font

def createDstFile(baseList, dstFileName, wb, ws): #resulList, dstFile
    print("\nBeginne jetzt mit dem schreiben des dst Files createDstFile()")
    
    for row in baseList: #Hängt jeden Eintrag in der resultList an das neu erzeugte sheet
        ws.append(row)

        # Formatiert die Überschriftenzeile
    for rows in ws.iter_rows(min_row=1, max_row=1):
            for cell in rows:
                cell.alignment = cell.alignment.copy(wrapText=True) # Zeilenumbruch
                cell.alignment = cell.alignment.copy(horizontal='left', vertical='center') # Ausrichtung
                cell.fill = PatternFill(start_color="2f48ff", end_color="2f48ff", fill_type = "solid") # Hintergrundfarbe
                cell.font = Font(size=14, color="FFFFFF") # Schriftfarbe
    
        # Formatiert die Restlichen Zellen
    for rows in ws.iter_rows(min_row=2):
            for cell in rows:
                cell.alignment = cell.alignment.copy(horizontal='left', vertical='center') # Ausrichtung
          
        # formatiert die Breite aller Spalten
    ws.column_dimensions['A'].width = 10 # Zir number
    ws.column_dimensions['B'].width = 14 # Repair number
    ws.column_dimensions['C'].width = 13 # Property
    ws.column_dimensions['D'].width = 12 # Date of creation
    ws.column_dimensions['E'].width = 12 # Creation Day
    ws.column_dimensions['F'].width = 13 # Creation Month 
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 21
    ws.column_dimensions['I'].width = 14 # Status Reference
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 12 # Supplier number
    ws.column_dimensions['L'].width = 42
    ws.column_dimensions['M'].width = 21
    ws.column_dimensions['N'].width = 44
    ws.column_dimensions['O'].width = 31
    ws.column_dimensions['P'].width = 27
    ws.column_dimensions['Q'].width = 17
    ws.column_dimensions['R'].width = 40
    
    ws.auto_filter.ref = ws.dimensions          #Setzt den Autofilter über alle Spalten

    try:
        wb.save(dstFileName) #Speichert das File
    except PermissionError:
        print("Die Zieldatei kann nicht geschrieben werden weil sie noch in einem anderem Programm geöffnet ist \n<< Enter >> beendet das Programm")
    else:
        wb.close()
        print("Das schreiben des Ausgabefiles ist abgeschlossen. Fürs Erste sind wir fertig")
    
           