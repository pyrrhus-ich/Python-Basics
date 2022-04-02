import os
from openpyxl import Workbook, load_workbook

workDir = os.getcwd()           
srcFld=workDir + "\\Anwendungen\\Zir_File_RepNrAbschneidenUndDatumUmwandeln\\srcZir\\"    
srcFile = srcFld + os.listdir(srcFld)[0]  
srcFileName=os.listdir(srcFld)[0]  
dstFile = workDir +"\\Anwendungen\\Zir_File_RepNrAbschneidenUndDatumUmwandeln\\dstZir\\changed-"+ srcFileName

# Wichtig für das korrekte Einlesen des SourceFiles
wb=load_workbook(filename=srcFile)
ws=wb.worksheets[0]

"""
Wichtig die folgenden Variablen bezeichnet den Index der Spalte in der das "Date of creation" und die Zir Reparatur
auftragsnummer ("M199-12345-R") steht
Index fängt bei 0 an zu zählen. Also ist 2 die 3. Spalte
"""
indRepairNumber= 1
indDateOfCreation = 3

# Variable die alls Werte aus dem sourceFile aufnimmt
val = []

# Wichtig für das erstellen des neuen Files
wbDst = Workbook()       # erzeugt Workbook Objekt mit einem Sheet
wsDst = wbDst.active     # aktiviert das erste sheet
wsDst.title="ZIR_ALL"
wsDst.sheet_properties.tabColor = "FF8800"